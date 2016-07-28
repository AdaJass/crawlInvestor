import aiohttp
import asyncio
import config
import processData as pd
import sys


from openpyxl import Workbook #写入excel使用(支持07)
from openpyxl.writer.excel import ExcelWriter

async def fetchData(callback = pd.processData):
    #set request url and parameters here or you can pass from outside. 
    
    #use s.** request a webside will keep-alive the connection automaticaly,
    #so you can set multi request here without close the connection 
    #while in the same domain.
    #i.e. 
    #await s.get('***/page1')
    #await s.get('***/page2')
    ######################################################################## 
    cookies = {
        'gr_user_id' : '87b14305-11d0-48b6-bd8f-4d243cabe01d',
        'pgv_pvi' : '9987699712',
        'identity' : 'jass.ada%40qq.com',
        'remember_code' : '5QE6q7WHRi',
        'acw_tc' : 'AQAAAJLiJ0xmFwMAZDvwcrBeFZnFtCMH',
        'session': '93809e53d60b2986d23789a699e11244ea55e280',
        'pgv_si' : 's1540547584',
        '_gat' :'1',
        'gr_session_id_eee5a46c52000d401f969f4535bdaa78': '0d721d63-71f2-4a3d-ac45-4d9d12ddb9fc',
        '_ga' : 'GA1.2.125227903.1457342767',
        'Hm_lvt_1c587ad486cdb6b962e94fc2002edf89' : '1469520678,1469522213,1469675226',
        'Hm_lpvt_1c587ad486cdb6b962e94fc2002edf89' : '1469675235'
    }

    conn = aiohttp.TCPConnector(limit=config.REQ_AMOUNTS)    
    s = aiohttp.ClientSession(headers = config.HEADERS, cookies=cookies, connector=conn)   

    url = 'https://www.itjuzi.com/investfirm/'
    AllData=[]
    pages=1
    r='text'
    while  pages<=6484:
        coroutines = [s.get(url+str(pages+i)) for i in range(10)]                
        for coroutine in asyncio.as_completed(coroutines):             
            try:
                r = await coroutine
            except Exception:
                await asyncio.sleep(2)
                r = await coroutine
            if not r:
                continue
            data =  await r.text(encoding='utf-8')
            AllData.append(await callback(data, s))

        pages=pages+10
        # await asyncio.sleep(1)
        

################################################################
################################################################

    filename = 'InstitutionVCDB.xlsx'
    sheetname = 'Institution'
    wb = Workbook()                            #新建一个文件
    wb_writer = ExcelWriter(workbook = wb)     #用来写入文件
    ws = wb.worksheets[0]                      #新建一个sheet
    ws.title = sheetname                       #定义sheet名称
        
    for i, v in enumerate(AllData):      #表示遍历行数(最后一个数字循环不到)
        ws.cell(row=i+1,column=1).value= v['title'] 
        ws.cell(row=i+1,column=2).value= v['descript']
        tem=v['investField'] 
        for num in range(6):
            if num >= len(tem):
                break
            ws.cell(row=i+1,column=3+num).value = tem[num]

        tem=v['investTurn'] 
        for num in range(4):
            if num >= len(tem):
                break
            ws.cell(row=i+1,column=9+num).value = tem[num]

        ws.cell(row=i+1,column=13).value= v['url']
        ws.cell(row=i+1,column=14).value= v['phone']
        ws.cell(row=i+1,column=15).value= v['email']
        ws.cell(row=i+1,column=16).value= v['address']
        
    wb_writer.save(filename = filename)        #保存写入文件


if __name__ == '__main__':    
    loop = asyncio.get_event_loop()    
    tasks = [fetchData(pd.processData)]    
    loop.run_until_complete(asyncio.wait(tasks))
    loop.close() 
